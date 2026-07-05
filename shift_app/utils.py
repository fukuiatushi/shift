from .models import Shift, Staff
import openpyxl
from datetime import datetime
import re

def import_shift_excel(file_path, shiftfile_title):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active

    # ★ Excel A1 から勤務月を取得
    title_cell = ws.cell(row=1, column=1).value

    year = None
    month = None

    # 西暦（例：2026年8月勤務表）
    m = re.search(r"(\d{4})年(\d+)月", title_cell)
    if m:
        year = int(m.group(1))
        month = int(m.group(2))

    # 和暦（例：令和8年8月勤務表）
    m2 = re.search(r"令和(\d+)年(\d+)月", title_cell)
    if m2:
        year = int(m2.group(1)) + 2018
        month = int(m2.group(2))

    # 月だけ（例：8月勤務表）
    m3 = re.search(r"(\d+)月", title_cell)
    if m3 and not month:
        month = int(m3.group(1))
        year = datetime.now().year

    if not year or not month:
        raise ValueError("Excel A1 に '8月勤務表' のように月が入っていません。")

    # ★ 惇さんの勤務表は構造が固定 → 職員名は必ず 4 行目から
    staff_start_row = 4

    # ★ 日付行（CSV で確認済み）
    date_row = 2

    # ★ 曜日行（CSV で確認済み）
    weekday_row = 3

    # ★ 行事行（CSV で確認済み → 6行目）
    event_row = 6

    # ★ 職員名と勤務コードを読み取る
    for row in range(staff_start_row, ws.max_row + 1):
        staff_name = ws.cell(row=row, column=1).value
        if not staff_name:
            continue

        # 職員名を登録
        staff_obj, _ = Staff.objects.get_or_create(name=staff_name)

        # ★ 日付は B列（2列目）から横に並ぶ
        for col in range(2, ws.max_column + 1):

            date_value = ws.cell(row=date_row, column=col).value
            weekday_value = ws.cell(row=weekday_row, column=col).value
            code = ws.cell(row=row, column=col).value
            event_value = ws.cell(row=event_row, column=col).value

            # 日付が空ならスキップ
            if not date_value:
                continue

            # ★ Excel の日付は「標準」にしてあるので int() が使える
            try:
                day = int(date_value)
                date_obj = datetime(year, month, day).date()
            except:
                continue

            # ★ DB に登録
            Shift.objects.update_or_create(
                staff=staff_obj,
                date=date_obj,
                defaults={
                    "code": code,
                    "weekday": weekday_value,
                    "event": event_value
                }
            )

    return shiftfile_title


